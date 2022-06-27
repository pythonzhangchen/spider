# -*- coding: utf-8 -*-
# @Time : 2022/6/27 16:00
# @Author : chen.zhang
# @File : DB_excel.py
from openpyxl import load_workbook, Workbook

# 创建心excel文档
wb = Workbook()
# 新创建一个sheet
ws = wb.create_sheet('myDATA')
# 使用默认sheet
# ws = wb.active

ws.append(['name','年龄'])

wb.save('111.xlsx')

# 开启excel 写入处理后的数据
result_list = [
    ['正正', '28'],
    ['培培', '35'],
    ['宁夫', '29'],
]

wb = load_workbook('111.xlsx')
ws = wb['myDATA']

for row in result_list:
    ws.append(row)
wb.save(filename='111.xlsx')

for x in ws.iter_rows(min_row=1):
    temp_row=[]
    for y in x:
        temp_row.append(y.value)
    print(temp_row)